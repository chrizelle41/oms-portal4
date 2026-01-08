import os
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from tqdm import tqdm

BASE_DIR = Path(__file__).resolve().parents[1]
INPUT_ROOT = BASE_DIR / "data" / "Input_Documents"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True, parents=True)


def extract_text_from_pdf(path: Path) -> str:
    try:
        reader = PdfReader(str(path))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages).strip()
    except Exception as e:
        print(f"[WARN] PDF read failed for {path}: {e}")
        return ""


def extract_text_from_docx(path: Path) -> str:
    try:
        doc = DocxDocument(str(path))
        return "\n".join(p.text for p in doc.paragraphs).strip()
    except Exception as e:
        print(f"[WARN] DOCX read failed for {path}: {e}")
        return ""


def extract_text_from_xlsx(path: Path) -> str:
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        texts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(v) for v in row if v is not None]
                if row_vals:
                    texts.append(" | ".join(row_vals))
        return "\n".join(texts).strip()
    except Exception as e:
        print(f"[WARN] XLSX read failed for {path}: {e}")
        return ""


def extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_text_from_pdf(path)
    elif ext in (".docx",):
        return extract_text_from_docx(path)
    elif ext in (".xlsx", ".xlsm"):
        return extract_text_from_xlsx(path)
    else:
        # For now we ignore other formats; can extend later
        return ""


def main():
    records = []
    text_records = []

    if not INPUT_ROOT.exists():
        raise SystemExit(f"Input root does not exist: {INPUT_ROOT}")

    print(f"Scanning {INPUT_ROOT} ...")

    file_paths = []
    for root, _, files in os.walk(INPUT_ROOT):
        for f in files:
            file_paths.append(Path(root) / f)

    for path in tqdm(file_paths, desc="Processing files"):
        rel = path.relative_to(INPUT_ROOT)

        # building = top-level folder (Botanical Place / Edocs / etc.)
        parts = rel.parts
        building = parts[0] if len(parts) > 1 else "UNKNOWN"

        stat = path.stat()
        size_bytes = stat.st_size
        modified_time = datetime.fromtimestamp(stat.st_mtime).isoformat()

        ext = path.suffix.lower()

        # Basic metadata 
        record = {
            "document_id": str(rel).replace(os.sep, "/"),
            "building": building,
            "relative_path": str(rel).replace(os.sep, "/"),
            "filename": path.name,
            "extension": ext,
            "size_bytes": size_bytes,
            "last_modified": modified_time,
        }

        # Try to get some text (for AI & search later)
        text = ""
        if ext in [".pdf", ".docx", ".xlsx", ".xlsm"]:
            text = extract_text(path)

        record["has_text"] = bool(text)

        records.append(record)

        if text:
            text_records.append({
                "document_id": record["document_id"],
                "building": building,
                "relative_path": record["relative_path"],
                "text": text[:200000]  # safety cap
            })

    # Save metadata CSV
    meta_df = pd.DataFrame(records)
    meta_path = OUTPUT_DIR / "documents_metadata.csv"
    meta_df.to_csv(meta_path, index=False)
    print(f"Saved metadata to {meta_path}")

    # Save text JSONL
    text_path = OUTPUT_DIR / "documents_text.jsonl"
    with text_path.open("w", encoding="utf-8") as f:
        for rec in text_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    print(f"Saved text chunks to {text_path}")


if __name__ == "__main__":
    main()